#!/usr/bin/env python
# -*- coding:utf-8 -*- 

import json
import os
import re
import openpyxl
from xlwt import Workbook

person_data = {}

#Judge the commit msg.
def judge_commit_msg_hook(subject):
    re1 = '.*?'  # Non-greedy match on filler
    re2 = '(Rootcause)'  # Square Braces 1
    re3 = '.*?'  # Non-greedy match on filler
    re4 = '(Solution)'  # Square Braces 2
    re5 = '.*?'  # Non-greedy match on filler
    re6 = '(Dependence)'  # Square Braces 3
    re7 = '.*?'
    rg = re.compile(re1 + re2 + re3 + re4 + re5 + re6 + re7, re.IGNORECASE | re.DOTALL)
    m = rg.search(subject)
    print('subject is ', subject)
    if m:
        print('XXXXXXXXXXXXXXXXXXXXXXXXX')
        return True
    return False

#Parse commits in each json file.
def parse_commits(full_file_path, json_data):
    commits_data =json_data['commits']
    commits_number = len(commits_data)
    print('commits number is ', commits_number)
    print(commits_data)

    #Parse the commits one bye one.
    for commit in commits_data:
        # First,parse the commit owner's name.
        #TODO: need to implement the all users in commits. Please refer to the GerritStats codes for align.
        owner_info = commit['owner']
        owner_name = owner_info['name'] + ' (' + owner_info['username'] + ')'
        print('owner_name is ', owner_name)

        #Second,judge wether the commit-hook used in this commit.
        subject = commit['subject']
        is_commit_msg_hook_used = judge_commit_msg_hook(subject)

        #Third, add/modify the record to the final person_data
        result = 0
        if not is_commit_msg_hook_used:
            result = 1
        if owner_name in person_data:
            hit_number = person_data[owner_name] + result
            person_data[owner_name] = hit_number
        else:
            person_data[owner_name] = result

    return

#Parse the json files in given folders.
def parse_json_files(json_file_path):
    files = os.listdir(json_file_path)
    print(len(files))
    for file in files:
        if not os.path.isdir(file):
            full_file_path = json_file_path + '\\'+ file
            print(full_file_path)
            with open(full_file_path, 'r', encoding="gbk") as f:
                json_data = json.load(f)
                parse_commits(full_file_path, json_data)

    return

#write the person_data to excel file.
def write_excel_file(excel_file_path):
    book = Workbook()
    sheet1 = book.add_sheet('subject_check')
    i = 0
    for (k, v) in person_data.items():
        sheet1.write(i, 0, k)
        sheet1.write(i, 1, v)
        i += 1
    book.save(excel_file_path)
    
    return

# Rewrite the Weekly Statistics of Geely CV Code Review Activities excel to add msg-hook info.
def rewrite_excel_file(excel_file_path):
    old_excel = openpyxl.load_workbook(excel_file_path)
    new_sheet = old_excel.get_active_sheet()
    for i in range(3, new_sheet.max_row, 1):
        cell_value = new_sheet.cell(row = i, column = 2).value
        if cell_value in person_data:
            new_sheet.cell(row = i, column = 5, value = person_data[cell_value])
        else:
            new_sheet.cell(row = i, column = 5, value = 0)
    old_excel.save(excel_file_path)

    return

if __name__ == '__main__':
    json_file_path = 'Z:\gerritstats\gerrit_out'
    excel_file_path = 'Z:\gerritstats\out-html\subject_check.xls'
    new_excel_file_path = 'D:\Project\Geely CV\Latest_Week_Data.xlsx'

    parse_json_files(json_file_path)
    print(person_data)
    print('len of person_data is ', len(person_data))
    #write_excel_file(excel_file_path)
    rewrite_excel_file(new_excel_file_path)